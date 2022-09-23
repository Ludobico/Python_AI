from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws = wb.create_sheet('diary',0) # 0이면 첫번째 위치에 삽입, 생략시 마지막 위치에 삽입

data = [('홍길동',80,70,90),('이기자',90,60,80),('강기자',80,80,70)]

r = 1 #첫번째 행이 1부터 시작
c = 1 #첫번째 열이 1부터 시작

for irum,kor,eng,math in data:
    ws.cell(row=r,column=c).value = irum
    ws.cell(row=r, column=c+1).value = kor
    ws.cell(row=r, column=c+2).value = eng
    ws.cell(row=r, column=c+3).value = math
    r += 1

"""
for irum, kor, eng, math in data:
    ws['A' + str(r)].value = irum
    ws['B' + str(r)].value = kor
    ws['C' + str(r)].value = eng
    ws['D' + str(r)].value = math
    r += 1
    
columnchar = 'A'
for irum, kor, eng, math in data:
    ws[columnchar + str(r)].value = irum
    ws[chr(ord(columnchar)+1) + str(r)].value = kor
    ws[chr(ord(columnchar)+2) + str(r)].value = eng
    ws[chr(ord(columnchar)+3) + str(r)].value = math
    r += 1
    
columnchar = 'A'
for val in data:
    for i in range(0,4):
        ws[chr(ord(columnchar)+i) + str(r).value = val[i]
    r += 1
"""

ws.cell(row=r, column=1).value = '합계'
ws.cell(row=r, column=2).value = '=sum(B1:B3)'
ws.cell(row=r, column=3).value = '=sum(C1:C3)'
ws.cell(row=r, column=4).value = '=sum(D1:D3)'

wb.save('openpyxl2.xlsx')
wb.close()
